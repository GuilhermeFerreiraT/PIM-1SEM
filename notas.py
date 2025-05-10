import json
import os  # Import necessário para localizar a pasta de Downloads
from config import console, sessao
from utils import carregar_progresso, carregar_cursos, carregar_usuarios
from collections import defaultdict
import pandas as pd
import numpy as np  # Substituir statistics por numpy
from openpyxl import Workbook

USUARIOS_JSON = "NP1OFC/JSON/usuarios.json"

def salvar_usuarios(usuarios):
    # Salva os usuários no arquivo JSON.
    with open(USUARIOS_JSON, 'w', encoding='utf-8') as arquivo:
        json.dump(usuarios, arquivo, indent=4, ensure_ascii=False)

def atualizar_classe_usuario(usuario_email):
    # Atualiza a classe do usuário com base no progresso.
    progresso = carregar_progresso()
    cursos = carregar_cursos()
    usuarios = carregar_usuarios()

    usuario = next((u for u in usuarios if u['email'] == usuario_email), None)
    if not usuario or usuario['perfil'] != 'aluno':
        return

    # Verificar se o usuário completou todos os cursos da classe atual
    cursos_classe_atual = [curso for curso in cursos if curso['classe'] == usuario['classe']]
    cursos_concluidos = {registro['curso'] for registro in progresso if registro['usuario'] == usuario_email}

    if all(curso['titulo'] in cursos_concluidos for curso in cursos_classe_atual):
        usuario['classe'] += 1  # Avança para a próxima classe
        salvar_usuarios(usuarios)
        console.print(f"[green]Usuário {usuario['nome']} avançou para a classe {usuario['classe']}![/green]")

def adicionar_nota_maxima_cursos(df_cursos, df_questionarios):
    """
    Adiciona a nota máxima aos cursos com base na quantidade de questionários vinculados.
    """
    # Calcula a quantidade de questionários por curso
    questionarios_por_curso = df_questionarios.groupby('curso')['id'].count().reset_index()
    questionarios_por_curso.columns = ['titulo', 'nota_maxima']

    # Adiciona a nota máxima ao DataFrame de cursos
    df_cursos = df_cursos.merge(questionarios_por_curso, on='titulo', how='left')
    df_cursos['nota_maxima'] = df_cursos['nota_maxima'].fillna(0).astype(int)

    return df_cursos

def gerar_relatorio_excel():
    from utils import carregar_questionarios
    from openpyxl.styles import PatternFill, Alignment
    from openpyxl import load_workbook

    # Gera um relatório Excel com as notas.
    progresso = carregar_progresso()
    usuarios = carregar_usuarios()
    cursos = carregar_cursos()
    questionarios = carregar_questionarios()

    if not progresso:
        console.print("[yellow]Nenhum progresso registrado até o momento.[/yellow]")
        return

    # Preparar DataFrames
    df_progresso = pd.DataFrame(progresso)
    df_usuarios = pd.DataFrame(usuarios)
    df_cursos = pd.DataFrame(cursos)
    df_questionarios = pd.DataFrame(questionarios)

    # Aba 1: Notas por Curso
    notas_curso = (
        df_progresso.groupby(['usuario', 'curso'])['pontos']
        .sum()
        .reset_index()
        .merge(df_usuarios[['email', 'nome']], left_on='usuario', right_on='email')
        .merge(df_cursos[['titulo']], left_on='curso', right_on='titulo')
    )
    notas_curso['nota_maxima'] = notas_curso['curso'].apply(
        lambda curso: len(df_questionarios[df_questionarios['curso'] == curso])
    )
    notas_curso['situacao'] = notas_curso.apply(
        lambda x: 'Aprovado' if x['pontos'] >= 0.6 * x['nota_maxima'] else 'Reprovado', axis=1
    )
    notas_curso = notas_curso[['nome', 'curso', 'pontos', 'situacao']]
    notas_curso.columns = ['Nome do Aluno', 'Curso', 'Nota', 'Situação']

    # Aba 2: Notas por Módulo
    notas_modulo = (
        df_progresso.merge(df_usuarios[['email', 'nome']], left_on='usuario', right_on='email')
        .merge(df_cursos[['titulo']], left_on='curso', right_on='titulo')
    )
    notas_modulo = notas_modulo[['nome', 'curso', 'modulo', 'pontos']]
    notas_modulo.columns = ['Nome do Aluno', 'Curso', 'Módulo', 'Nota do Módulo']

    # Aba 3: Progresso por Módulo
    progresso_modulo = (
        df_progresso.groupby(['usuario', 'curso'])['modulo']
        .nunique()
        .reset_index()
        .merge(df_usuarios[['email', 'nome']], left_on='usuario', right_on='email')
        .merge(df_cursos[['titulo', 'modulos']], left_on='curso', right_on='titulo')
    )
    progresso_modulo['total_modulos'] = progresso_modulo['modulos'].apply(len)
    progresso_modulo['situacao'] = progresso_modulo.apply(
        lambda x: 'Completo' if x['modulo'] == x['total_modulos'] else
                  'Em andamento' if x['modulo'] > 0 else 'Incompleto', axis=1
    )
    progresso_modulo = progresso_modulo[['nome', 'curso', 'total_modulos', 'modulo', 'situacao']]
    progresso_modulo.columns = ['Nome do Aluno', 'Curso', 'Número total de Módulos', 'Número de Módulos Completos', 'Situação']

    # Aba 4: Estatísticas por Módulo
    estatisticas_modulo = (
        df_progresso.groupby(['curso', 'modulo'])['pontos']
        .agg(mean='mean', moda=lambda x: x.mode().iloc[0] if not x.mode().empty else np.nan, mediana='median')
        .reset_index()
    )
    estatisticas_modulo.columns = ['Curso', 'Módulo', 'Média das Notas', 'Moda das Notas', 'Mediana das Notas']

    # Aba 5: Estatísticas por Curso
    estatisticas_curso = (
        df_progresso.groupby('curso')['pontos']
        .agg(mean='mean', moda=lambda x: x.mode().iloc[0] if not x.mode().empty else np.nan, mediana='median')
        .reset_index()
    )
    estatisticas_curso.columns = ['Curso', 'Média das Notas', 'Moda das Notas', 'Mediana das Notas']

    # Criar arquivo Excel
    caminho_downloads = os.path.join(os.path.expanduser("~"), "Downloads")
    if not os.path.exists(caminho_downloads):
        console.print("[red]Erro: A pasta de Downloads não foi encontrada.[/red]")
        return

    caminho_arquivo = os.path.join(caminho_downloads, "relatorio_notas.xlsx")
    try:
        with pd.ExcelWriter(caminho_arquivo, engine='openpyxl') as writer:
            notas_curso.to_excel(writer, index=False, sheet_name="Notas por Curso")
            notas_modulo.to_excel(writer, index=False, sheet_name="Notas por Módulo")
            progresso_modulo.to_excel(writer, index=False, sheet_name="Progresso por Módulo")
            estatisticas_modulo.to_excel(writer, index=False, sheet_name="Estatísticas por Módulo")
            estatisticas_curso.to_excel(writer, index=False, sheet_name="Estatísticas por Curso")

        # Aplicar formatações no Excel
        workbook = load_workbook(caminho_arquivo)
        fills = {
            "header": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "green": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
            "red": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
            "yellow": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        }

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            # Ajustar largura das colunas
            for col in sheet.columns:
                max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
                sheet.column_dimensions[col[0].column_letter].width = max_length + 2

            # Aplicar cor ao cabeçalho
            for cell in sheet[1]:
                cell.fill = fills["header"]
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Formatação condicional para "Notas por Curso" e "Progresso por Módulo"
            if sheet_name == "Notas por Curso":
                for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=4, max_col=4):
                    for cell in row:
                        if cell.value == "Aprovado":
                            cell.fill = fills["green"]
                        elif cell.value == "Reprovado":
                            cell.fill = fills["red"]
            elif sheet_name == "Progresso por Módulo":
                for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=5, max_col=5):
                    for cell in row:
                        if cell.value == "Completo":
                            cell.fill = fills["green"]
                        elif cell.value == "Em andamento":
                            cell.fill = fills["yellow"]

        workbook.save(caminho_arquivo)
        console.print(f"[green]Relatório Excel gerado com sucesso: {caminho_arquivo}[/green]")
    except PermissionError:
        console.print("[red]Erro: O arquivo está aberto ou sem permissão para ser sobrescrito. Feche o arquivo e tente novamente.[/red]")

def input_numerico(mensagem, minimo=None, maximo=None):
    while True:
        try:
            valor = int(console.input(mensagem))
            if (minimo is not None and valor < minimo) or (maximo is not None and valor > maximo):
                console.print(f"[red]Por favor, insira um número entre {minimo} e {maximo}.[/red]")
                continue
            return valor
        except ValueError:
            console.print("[red]Entrada inválida. Por favor, insira um número válido.[/red]")
